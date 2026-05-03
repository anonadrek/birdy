"""Build species_list.yaml from VP11.pdf + IOC v14.1 xlsx with Wikidata Q-ID mapping."""

from __future__ import annotations

import asyncio
import json
import re
from pathlib import Path
from typing import Any

import aiohttp
import openpyxl
import pdfplumber
import yaml

from .models import (
    IocEntry,
    MappingFailure,
    SpeciesListEntry,
    Vp11Entry,
    VpStatus,
)

WIKIDATA_SPARQL = "https://query.wikidata.org/sparql"
USER_AGENT = "birdy-fetcher/0.1.0 (https://github.com/anonadrek/birdy)"

# VP11 column x-coordinates (verifierat med pdfplumber.extract_words på riktiga PDF:en)
VP11_X_STATUS_END = 145
VP11_X_SCI_END = 265
VP11_X_SWE_END = 355  # svenska kolumnen — vi extraherar inte (font-defekt)
VP11_X_ENG_END = 455
# notes: x >= 455

VALID_VP_STATUSES: set[VpStatus] = {"H", "h", "F", "R", "(H)"}

# Genus species (med valfri 3-ords subspecies) — strikt 2-3 ord lowercase efter Genus
SCI_NAME_RE = re.compile(r"^[A-Z][a-zA-Z\-']+\s+[a-z\-']+(?:\s+[a-z\-']+)?$")


_IOC_ORDER_RE = re.compile(r"^ORDER\s+([A-Z]+)$")
_IOC_FAMILY_RE = re.compile(r"^Family\s+([A-Z][a-zA-Z]+)$")


def parse_ioc(xlsx_path: Path) -> list[IocEntry]:
    """Parse IOC v14.1 Excel master list.

    IOC är hierarkiskt: rader har Rank ∈ {INFRACLASS, ORDER, Family, Genus, Species, ssp, Blank}.
    'Scientific Name' innehåller "ORDER X" / "Family Y" / Genus / binom beroende på rank.
    Vi traverserar och bygger upp current_order/current_family, och plockar bara Species-rader.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["14.1"] if "14.1" in wb.sheetnames else wb.active
    if ws is None:
        raise ValueError(f"No usable sheet in {xlsx_path}")
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows)]

    out: list[IocEntry] = []
    current_order: str | None = None
    current_family: str | None = None
    current_family_en: str | None = None

    for row in rows:
        d = dict(zip(headers, row, strict=False))
        rank = (str(d.get("Rank") or "")).strip()
        sci_field = (str(d.get("Scientific Name") or "")).strip()
        eng = (str(d.get("English name") or "")).strip()

        if rank == "ORDER":
            m = _IOC_ORDER_RE.match(sci_field)
            current_order = m.group(1).capitalize() if m else None
            continue
        if rank == "Family":
            m = _IOC_FAMILY_RE.match(sci_field)
            current_family = m.group(1) if m else None
            current_family_en = eng or None
            continue
        if rank != "Species":
            continue
        if not sci_field or not current_order or not current_family:
            continue

        out.append(
            IocEntry(
                scientific_name=sci_field,
                family=current_family,
                family_en=current_family_en,
                ioc_order=current_order,
                common_en=eng,
            )
        )
    return out


def parse_vp11(pdf_path: Path) -> list[Vp11Entry]:
    """Parse VP11.pdf with pdfplumber column coordinates.

    Svenska namn extraheras inte — PDF:en har defekt ToUnicode-mapping för åäö-glyfer.
    Svenska namn hämtas från Wikidata i Task 4. Här tar vi status/sci/eng/notes.
    """
    out: list[Vp11Entry] = []
    current_order: str | None = None
    current_family: str | None = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            # Gruppera per rad med snap till 2pt
            lines: dict[int, list[dict[str, Any]]] = {}
            for w in words:
                y_key = round(float(w["top"]) / 2) * 2
                lines.setdefault(y_key, []).append(w)
            for y in sorted(lines):
                row_words = sorted(lines[y], key=lambda w: float(str(w["x0"])))
                cols = _split_vp11_columns(row_words)
                sci = cols["sci"]

                # Hierarki-rader — dessa kommer i "sci"-spalten (x ~147)
                if sci.startswith("Ordning "):
                    m = re.match(r"^Ordning\s+([A-Z]+)", sci)
                    # Capitalize matches IOC's "Passeriformes" output for cross-source consistency
                    current_order = m.group(1).capitalize() if m else None
                    continue
                if sci.startswith("Familj "):
                    m = re.match(r"^Familj\s+([A-Z][a-zA-Z\-']+)", sci)
                    current_family = m.group(1) if m else None
                    continue

                # Skippa header / fotnot
                if not sci or "VETENSKAPLIGT" in sci or "Scientific" in sci:
                    continue

                status = cols["status"]
                if status not in VALID_VP_STATUSES:
                    continue
                if not SCI_NAME_RE.match(sci):
                    continue
                if current_order is None or current_family is None:
                    continue

                out.append(
                    Vp11Entry(
                        status=status,
                        ioc_order=current_order,
                        family=current_family,
                        scientific_name=sci,
                        common_en=cols["eng"],
                        notes=cols["notes"],
                    )
                )
    return out


def _split_vp11_columns(row_words: list[dict[str, Any]]) -> dict[str, str]:
    # pdfplumber emits no table structure — we rebuild columns from x-coordinates.
    buckets: dict[str, list[str]] = {"status": [], "sci": [], "eng": [], "notes": []}
    for w in row_words:
        x = float(w["x0"])
        text = str(w["text"])
        if x < VP11_X_STATUS_END:
            buckets["status"].append(text)
        elif x < VP11_X_SCI_END:
            buckets["sci"].append(text)
        elif x < VP11_X_SWE_END:
            pass  # svenska kolumnen — skippas (font-defekt på åäö)
        elif x < VP11_X_ENG_END:
            buckets["eng"].append(text)
        else:
            buckets["notes"].append(text)
    return {k: " ".join(v).strip() for k, v in buckets.items()}


def cross_check_with_ioc(
    vp11_entries: list[Vp11Entry], ioc_entries: list[IocEntry]
) -> tuple[list[Vp11Entry], list[MappingFailure]]:
    """Filtrera VP11-arter där vetenskapligt namn finns i IOC. Skillnader → failures."""
    ioc_by_sci = {e.scientific_name: e for e in ioc_entries}
    matched: list[Vp11Entry] = []
    failures: list[MappingFailure] = []
    for v in vp11_entries:
        if v.scientific_name in ioc_by_sci:
            matched.append(v)
        else:
            failures.append(
                MappingFailure(
                    scientific_name=v.scientific_name,
                    family=v.family,
                    common_en=v.common_en,
                    reason="VP11 sci_name not found in IOC v14.1 — taxonomy drift?",
                )
            )
    return matched, failures


def _build_sparql_query(scientific_names: list[str]) -> str:
    values = " ".join(f'"{n}"' for n in scientific_names)
    return f"""
    SELECT ?item ?scientificName WHERE {{
      VALUES ?scientificName {{ {values} }}
      ?item wdt:P225 ?scientificName .
      ?item wdt:P31/wdt:P279* wd:Q16521 .
    }}
    """


async def _run_sparql(query: str) -> str:
    """Run a SPARQL query against Wikidata. Override in tests."""
    async with (
        aiohttp.ClientSession(headers={"User-Agent": USER_AGENT}) as session,
        session.get(
            WIKIDATA_SPARQL,
            params={"query": query, "format": "json"},
            timeout=aiohttp.ClientTimeout(total=60),
        ) as response,
    ):
        response.raise_for_status()
        return await response.text()


async def map_to_wikidata(scientific_names: list[str]) -> dict[str, str]:
    """Return mapping from scientific_name → Q-ID. Missing names omitted."""
    if not scientific_names:
        return {}
    query = _build_sparql_query(scientific_names)
    raw = await _run_sparql(query)
    data = json.loads(raw)
    result: dict[str, str] = {}
    for binding in data["results"]["bindings"]:
        sn = binding["scientificName"]["value"]
        uri = binding["item"]["value"]
        q_id = uri.rsplit("/", 1)[-1]
        result[sn] = q_id
    return result


async def _resolve_wikidata_qids(
    scientific_names: list[str],
    *,
    batch_size: int,
    concurrency: int,
) -> dict[str, str]:
    """Run map_to_wikidata in concurrent batches to amortize SPARQL latency."""
    if not scientific_names:
        return {}
    sem = asyncio.Semaphore(concurrency)

    async def _bounded(batch: list[str]) -> dict[str, str]:
        async with sem:
            return await map_to_wikidata(batch)

    batches = [
        scientific_names[i : i + batch_size] for i in range(0, len(scientific_names), batch_size)
    ]
    results = await asyncio.gather(*(_bounded(b) for b in batches))
    out: dict[str, str] = {}
    for r in results:
        out.update(r)
    return out


def _merge_with_existing(
    pipeline_entries: list[SpeciesListEntry], existing_yaml: Path
) -> list[SpeciesListEntry]:
    # Manual additions (entries the user added after reviewing mapping_failures) win
    # over a re-run pipeline output for any name not produced by the pipeline.
    existing_data = yaml.safe_load(existing_yaml.read_text(encoding="utf-8")) or []
    pipeline_names = {e.scientific_name for e in pipeline_entries}
    manual_extras = [
        SpeciesListEntry(**d)
        for d in existing_data
        if d.get("scientific_name") not in pipeline_names
    ]
    return pipeline_entries + manual_extras


async def build_species_list(
    *,
    ioc_xlsx: Path,
    vp11_pdf: Path,
    filter_yaml: Path,
    out_list: Path,
    out_failures: Path,
    resume: bool = False,
    batch_size: int = 50,
    sparql_concurrency: int = 5,
) -> None:
    ioc = parse_ioc(ioc_xlsx)
    vp11 = parse_vp11(vp11_pdf)
    filter_data = yaml.safe_load(filter_yaml.read_text(encoding="utf-8"))
    include_statuses: set[str] = set(filter_data["include_statuses"])

    # Filter VP11 by status
    vp11_filtered = [v for v in vp11 if v.status in include_statuses]

    # Cross-check mot IOC
    matched, ioc_failures = cross_check_with_ioc(vp11_filtered, ioc)

    qid_map = await _resolve_wikidata_qids(
        [v.scientific_name for v in matched],
        batch_size=batch_size,
        concurrency=sparql_concurrency,
    )

    entries: list[SpeciesListEntry] = []
    wikidata_failures: list[MappingFailure] = []
    for v in matched:
        qid = qid_map.get(v.scientific_name)
        if qid is None:
            wikidata_failures.append(
                MappingFailure(
                    scientific_name=v.scientific_name,
                    family=v.family,
                    common_en=v.common_en,
                    reason="no Wikidata match for P225",
                )
            )
            continue
        entries.append(
            SpeciesListEntry(
                wikidata_id=qid,
                scientific_name=v.scientific_name,
                family=v.family,
                ioc_order=v.ioc_order,
                common_en=v.common_en,
                vp_status=v.status,
            )
        )

    if resume and out_list.exists():
        entries = _merge_with_existing(entries, out_list)

    # Auto-cleanup: any failure resolved by a manual addition disappears from failures.
    resolved_names = {e.scientific_name for e in entries}
    all_failures = [
        f for f in (ioc_failures + wikidata_failures) if f.scientific_name not in resolved_names
    ]

    out_list.parent.mkdir(parents=True, exist_ok=True)
    out_list.write_text(
        yaml.safe_dump(
            [e.model_dump(exclude_none=True) for e in entries],
            sort_keys=False,
            allow_unicode=True,
        ),
        encoding="utf-8",
    )
    out_failures.write_text(
        _FAILURES_HEADER
        + yaml.safe_dump(
            [f.model_dump() for f in all_failures],
            sort_keys=False,
            allow_unicode=True,
        ),
        encoding="utf-8",
    )


_FAILURES_HEADER = (
    "# Each entry below is a species the automatic mapper could not match.\n"
    "# To resolve a failure:\n"
    "#   1. Search wikidata.org for the scientific name → note the Q-ID\n"
    "#   2. Add an entry to species_list.yaml manually\n"
    "#      (with wikidata_id, scientific_name, family, ioc_order, common_en, vp_status)\n"
    "#   3. Run: uv run birdy-fetcher init --resume\n"
    "# --resume re-reads species_list.yaml, preserves your manual additions,\n"
    "# and removes resolved entries from this file automatically.\n"
    "---\n"
)


async def cli_init(
    *,
    sources_dir: Path,
    checklists_dir: Path,
    out_dir: Path,
    resume: bool = False,
) -> int:
    """Returns exit code: 0 if no failures, 1 if mapping_failures.yaml is non-empty."""
    out_list = out_dir / "species_list.yaml"
    out_failures = out_dir / "mapping_failures.yaml"
    if resume and not out_list.exists():
        raise RuntimeError("nothing to resume — species_list.yaml not found")

    await build_species_list(
        ioc_xlsx=sources_dir / "ioc-14.1.xlsx",
        vp11_pdf=sources_dir / "vp11.pdf",
        filter_yaml=checklists_dir / "vp11-filter.yaml",
        out_list=out_list,
        out_failures=out_failures,
        resume=resume,
    )

    failures = yaml.safe_load(out_failures.read_text(encoding="utf-8")) or []
    return 0 if not failures else 1
